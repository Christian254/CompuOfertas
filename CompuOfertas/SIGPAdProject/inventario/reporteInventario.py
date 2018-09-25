# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from datetime import datetime
from decimal import *
from openpyxl import Workbook
from openpyxl.styles import Alignment 
from openpyxl.styles import Border, Side
from django.db.models import Q
from django.contrib.contenttypes.models import ContentType
from .models import *
from django.http import HttpResponse

def descargarExcel():
	productos = Producto.objects.filter(estado=1)
	wb = Workbook()
	ws = wb.active
	ws['B1'] = 'Reporte de inventario'
	ws.merge_cells('B1:I1')
	ws.merge_cells('B2:I2')
	ws['B3']= 'categoria'
	ws['C3']= 'codigo producto'
	ws['D3']= 'nombre producto'
	ws['E3']= 'marca'
	ws['F3']= 'descripcion'
	ws['G3']= 'Existencia'
	ws['H3']= 'precio promedio de compra'
	ws['I3']= 'precio de venta'
	ws['B1'].alignment = Alignment(horizontal='center')
	ws.column_dimensions['B'].width = 10
	ws.column_dimensions['C'].width = 10
	ws.column_dimensions['D'].width = 15
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['F'].width = 20
	ws.column_dimensions['G'].width = 10
	ws.column_dimensions['H'].width = 25
	ws.column_dimensions['I'].width = 20

	cont = 4
	for p in productos:
		ws.cell(row = cont, column = 2).value = p.categoria.codigo
		ws.cell(row = cont, column = 3).value = p.codigo
		ws.cell(row = cont, column = 4).value = p.nombre
		ws.cell(row = cont, column = 5).value = p.marca
		ws.cell(row = cont, column = 6).value = p.descripcion
		ws.cell(row = cont, column = 7).value = p.inventario.existencia
		ws.cell(row = cont, column = 8).value = p.inventario.precio_promedio_compra
		ws.cell(row = cont, column = 9).value = p.inventario.precio_venta_producto
		cont += 1 

	for y in xrange(1,4):
		for x in xrange(2,10):
			ws.cell(row = y, column = x).border = Border(top = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin', color='FF000000'), bottom = Side(border_style='thin', color='FF000000'), left = Side(border_style='thin', color='FF000000'))

	contador = 4
	for y in productos:
		for x in xrange(2,10):
			ws.cell(row = contador, column = x).border = Border(top = Side(border_style='thin', color='FF000000'), right = Side(border_style='thin', color='FF000000'), bottom = Side(border_style='thin', color='FF000000'), left = Side(border_style='thin', color='FF000000'))
		contador += 1

	ws.protection.set_password('root1234')
	response = HttpResponse(content_type="application/ms-excel")
	response['Content-Disposition'] = 'attachment; filename=inventario.xlsx'
	wb.save(response)
	return response